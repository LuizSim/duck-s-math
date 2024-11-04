from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from pathlib import Path

app = Flask(__name__)
CORS(app)  # Permite requisições de todas as origens

responses = []

@app.route('/', methods=['POST'])
def save_response():
    if request.method == 'POST':
        data = request.json
        responses.append(data)
        return jsonify({"message": "Resposta registrada com sucesso"}), 200

@app.route('/', methods=['GET'])
def generate_excel():
    if responses:
        # Define o caminho do arquivo para salvar na pasta de downloads do usuário
        caminho_pasta_downloads = str(Path.home() / "Downloads")

        # Salva o arquivo Excel dentro da pasta de downloads
        caminho_arquivo = os.path.join(caminho_pasta_downloads, 'resultado_ducks_math.xlsx')
        
        try:
            # Criação da planilha usando openpyxl
            wb = Workbook()
            ws = wb.active
            ws.title = "Respostas"

            # Adiciona o cabeçalho
            ws.append(["Perguntas", "Resposta", "Resposta do jogador", "Acertos", "Score"])

            # Definir cores para acertos e erros
            green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            # Preencher a planilha com as respostas
            for response in responses:
                pergunta = response.get("question")
                resposta_correta = response.get("correct_answer")
                resposta_jogador = response.get("player_answer")
                acerto = "Correto" if resposta_correta == resposta_jogador else "Incorreto"
                score = response.get("score", 0)

                # Adiciona a linha na planilha
                ws.append([pergunta, resposta_correta, resposta_jogador, acerto, score])

                # Aplicar cor na célula de "Acertos"
                acertos_cell = ws.cell(row=ws.max_row, column=4)
                if acerto == "Correto":
                    acertos_cell.fill = green_fill
                else:
                    acertos_cell.fill = red_fill

            # Salvar a planilha
            wb.save(caminho_arquivo)

            return jsonify({"message": "Arquivo Excel gerado com sucesso", "file_path": caminho_arquivo}), 200

        except Exception as e:
            return jsonify({"message": f"Erro ao gerar o arquivo Excel: {str(e)}"}), 500

    return jsonify({"message": "Nenhuma resposta registrada"}), 400

if __name__ == '__main__':
    app.run(debug=True)